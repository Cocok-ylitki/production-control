from datetime import date, datetime
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from src.data.models import Batch


def _format_dt(d: datetime | None) -> str:
    if d is None:
        return "-"
    return d.strftime("%Y-%m-%d %H:%M:%S") if hasattr(d, "strftime") else str(d)


def _format_date(d: date | None) -> str:
    if d is None:
        return "-"
    return d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)


def generate_batch_report_excel(batch: "Batch") -> bytes:
    """Генерирует Excel-отчёт по партии: 3 листа (информация, продукция, статистика)."""
    wb = Workbook()
    # Лист 1: Информация о партии
    ws1 = wb.active
    ws1.title = "Информация о партии"
    info = [
        ("Номер партии", batch.batch_number),
        ("Дата партии", _format_date(batch.batch_date)),
        ("Статус", "Закрыта" if batch.is_closed else "Открыта"),
        ("Рабочий центр", batch.work_center.name if batch.work_center else "-"),
        ("Смена", batch.shift),
        ("Бригада", batch.team),
        ("Номенклатура", batch.nomenclature),
        ("Начало смены", _format_dt(batch.shift_start)),
        ("Окончание смены", _format_dt(batch.shift_end)),
    ]
    for row, (label, value) in enumerate(info, start=1):
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 30

    # Лист 2: Продукция
    ws2 = wb.create_sheet("Продукция")
    headers = ["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"]
    for col, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=col, value=h, font=Font(bold=True))
    for row, p in enumerate(batch.products, start=2):
        ws2.cell(row=row, column=1, value=p.id)
        ws2.cell(row=row, column=2, value=p.unique_code)
        ws2.cell(row=row, column=3, value="Да" if p.is_aggregated else "Нет")
        ws2.cell(row=row, column=4, value=_format_dt(p.aggregated_at))
    for c in range(1, 5):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # Лист 3: Статистика
    ws3 = wb.create_sheet("Статистика")
    total = len(batch.products)
    agg_count = sum(1 for p in batch.products if p.is_aggregated)
    rest = total - agg_count
    pct = (100 * agg_count / total) if total else 0
    ws3["A1"] = "Всего продукции:"
    ws3["B1"] = total
    ws3["A2"] = "Аггрегировано:"
    ws3["B2"] = agg_count
    ws3["A3"] = "Осталось:"
    ws3["B3"] = rest
    ws3["A4"] = "Процент выполнения:"
    ws3["B4"] = f"{pct:.0f}%"
    if batch.shift_start and batch.shift_end and agg_count:
        try:
            delta = (batch.shift_end - batch.shift_start).total_seconds() / 3600
            if delta > 0:
                ws3["A5"] = "Средняя скорость:"
                ws3["B5"] = f"{agg_count / delta:.0f} ед/час"
        except Exception:
            pass
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_batches_export_excel(batches: list["Batch"]) -> bytes:
    """Экспорт списка партий в Excel (один лист с таблицей)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Партии"
    headers = [
        "ID", "Номер партии", "Дата партии", "Статус", "Рабочий центр", "Смена", "Бригада",
        "Номенклатура", "Код ЕКН", "Описание", "Начало смены", "Окончание смены", "Создан",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h, font=Font(bold=True))
    for row, b in enumerate(batches, start=2):
        ws.cell(row=row, column=1, value=b.id)
        ws.cell(row=row, column=2, value=b.batch_number)
        ws.cell(row=row, column=3, value=_format_date(b.batch_date))
        ws.cell(row=row, column=4, value="Закрыта" if b.is_closed else "Открыта")
        ws.cell(row=row, column=5, value=b.work_center.name if b.work_center else "-")
        ws.cell(row=row, column=6, value=b.shift)
        ws.cell(row=row, column=7, value=b.team)
        ws.cell(row=row, column=8, value=b.nomenclature)
        ws.cell(row=row, column=9, value=b.ekn_code)
        ws.cell(row=row, column=10, value=b.task_description or "")
        ws.cell(row=row, column=11, value=_format_dt(b.shift_start))
        ws.cell(row=row, column=12, value=_format_dt(b.shift_end))
        ws.cell(row=row, column=13, value=_format_dt(b.created_at))
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_batches_export_csv(batches: list["Batch"]) -> bytes:
    """Экспорт списка партий в CSV."""
    import csv
    from io import StringIO
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "ID", "Номер партии", "Дата партии", "Статус", "Рабочий центр", "Смена", "Бригада",
        "Номенклатура", "Код ЕКН", "Описание", "Начало смены", "Окончание смены", "Создан",
    ])
    for b in batches:
        writer.writerow([
            b.id, b.batch_number, _format_date(b.batch_date),
            "Закрыта" if b.is_closed else "Открыта",
            b.work_center.name if b.work_center else "-",
            b.shift, b.team, b.nomenclature, b.ekn_code, b.task_description or "",
            _format_dt(b.shift_start), _format_dt(b.shift_end), _format_dt(b.created_at),
        ])
    return buf.getvalue().encode("utf-8-sig")
