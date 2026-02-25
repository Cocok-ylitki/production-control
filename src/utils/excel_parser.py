from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


# Ожидаемые заголовки (русские) и маппинг в ключи для создания партии
HEADER_MAP = {
    "номерпартии": "НомерПартии",
    "датапартии": "ДатаПартии",
    "номенклатура": "Номенклатура",
    "рабочийцентр": "РабочийЦентр",
    "смена": "Смена",
    "бригада": "Бригада",
    "кодекн": "КодЕКН",
    "идентификаторрц": "ИдентификаторРЦ",
    "представлениезаданиянасмену": "ПредставлениеЗаданияНаСмену",
    "статусзакрытия": "СтатусЗакрытия",
    "датавремяначаласмены": "ДатаВремяНачалаСмены",
    "датавремяокончаниясмены": "ДатаВремяОкончанияСмены",
}


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "")


def _parse_cell(value: Any, header_key: str) -> Any:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if header_key == "НомерПартии":
        return int(value) if not isinstance(value, int) else value
    if header_key == "ДатаПартии":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            return datetime.strptime(value[:10], "%Y-%m-%d").date()
        return value
    if header_key in ("ДатаВремяНачалаСмены", "ДатаВремяОкончанияСмены"):
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        return value
    if header_key == "СтатусЗакрытия":
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in ("1", "true", "да", "yes")
        return bool(value)
    return value if not isinstance(value, str) else value.strip()


def parse_batches_excel(data: bytes) -> list[tuple[dict, int]]:
    """
    Парсит Excel с партиями. Первая строка — заголовки.
    Возвращает список (row_dict, row_index) для валидных строк.
    row_index 1-based для сообщений об ошибках.
    """
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    col_index = {_normalize_header(str(h)): i for i, h in enumerate(header_row)}
    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_dict = {}
        for norm_name, canonical in HEADER_MAP.items():
            i = col_index.get(norm_name)
            if i is None:
                continue
            if i < len(row):
                val = _parse_cell(row[i], canonical)
                if val is not None or canonical == "СтатусЗакрытия":
                    row_dict[canonical] = val
        if row_dict:
            result.append((row_dict, row_idx))
    return result
